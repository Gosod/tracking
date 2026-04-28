"""
excel_parser.py — парсинг Excel-файла из 1С.

Структура файла:
  Строка 1: "Заказ №" в A1, номер заказа рядом (B1 или C1)
  Строка 2: заголовки колонок
  Строка 3+: данные позиций (могут быть мёрджнутые ячейки)

Нам нужны:
  Колонка 4 (D) — Обозначение
  Колонка 5 (E) — Наименование
  Колонка 6 (F) — Кол-во
"""

import openpyxl


def _get_cell_value(ws, row, col):
    """
    Получить значение ячейки с учётом мёрджнутых диапазонов.
    Если ячейка является частью мёрджа — берём значение из верхнего левого угла.
    """
    cell = ws.cell(row=row, column=col)
    
    if cell.value is not None:
        return cell.value
    
    # Ищем, является ли эта ячейка частью мёрджнутого диапазона
    for merge_range in ws.merged_cells.ranges:
        if (merge_range.min_row <= row <= merge_range.max_row and
                merge_range.min_col <= col <= merge_range.max_col):
            # Берём значение из верхнего левого угла мёрджа
            return ws.cell(merge_range.min_row, merge_range.min_col).value
    
    return None


def parse_excel(filepath):
    """
    Парсит Excel-файл из 1С.
    
    Возвращает dict:
      {
        "order_number": "ОЗТ-25-00787",
        "positions": [
          {"pos_number": 1, "designation": "PMA.IGC...", "name": "Борт сварной", "qty": 1.0},
          ...
        ]
      }
    
    Бросает ValueError если файл не удаётся распарсить.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        raise ValueError(f"Не удалось открыть файл: {e}")
    
    ws = wb.active
    
    # ── Ищем номер заказа в первой строке ────────────────────────────────────
    order_number = None
    
    # Ищем ячейку с текстом "Заказ" в первых 3 строках
    for row in range(1, 4):
        for col in range(1, 10):
            val = ws.cell(row=row, column=col).value
            if val and "заказ" in str(val).lower():
                # Номер заказа — в следующих ячейках той же строки
                for next_col in range(col + 1, col + 6):
                    next_val = _get_cell_value(ws, row, next_col)
                    if next_val and str(next_val).strip():
                        order_number = str(next_val).strip()
                        break
                if order_number:
                    break
        if order_number:
            break
    
    if not order_number:
        # Запасной вариант — просто берём значение из B1 или C1
        for col in range(2, 8):
            val = _get_cell_value(ws, 1, col)
            if val and str(val).strip():
                order_number = str(val).strip()
                break
    
    if not order_number:
        raise ValueError("Не удалось найти номер заказа в файле (строка 1)")
    
    # ── Ищем строку с заголовками ─────────────────────────────────────────────
    # Заголовки содержат "Обозначение", "Наименование" и т.д.
    header_row = None
    for row in range(1, 10):
        for col in range(1, 10):
            val = ws.cell(row=row, column=col).value
            if val and "обозначение" in str(val).lower():
                header_row = row
                break
        if header_row:
            break
    
    if not header_row:
        # Если заголовки не найдены — считаем что данные с 3-й строки
        header_row = 2
    
    # ── Определяем индексы нужных колонок по заголовкам ──────────────────────
    col_pos = None          # Поз.
    col_designation = None  # Обозначение
    col_name = None         # Наименование
    col_qty = None          # Кол-во
    
    for col in range(1, 20):
        val = ws.cell(row=header_row, column=col).value
        if not val:
            continue
        val_lower = str(val).lower().strip()
        if "поз" in val_lower:
            col_pos = col
        elif "обозначение" in val_lower:
            col_designation = col
        elif "наименование" in val_lower:
            col_name = col
        elif "кол" in val_lower:
            col_qty = col
    
    # Если автоопределение не сработало — используем фиксированные номера (из спеки)
    if not col_designation:
        col_designation = 4
    if not col_name:
        col_name = 5
    if not col_qty:
        col_qty = 6
    if not col_pos:
        col_pos = 1
    
    # ── Читаем позиции ────────────────────────────────────────────────────────
    positions = []
    data_start_row = header_row + 1
    
    for row in range(data_start_row, ws.max_row + 1):
        designation = _get_cell_value(ws, row, col_designation)
        name = _get_cell_value(ws, row, col_name)
        qty_raw = _get_cell_value(ws, row, col_qty)
        pos_raw = _get_cell_value(ws, row, col_pos)
        
        # Пропускаем пустые строки
        if not designation and not name:
            continue
        
        # Пропускаем строки где designation = "Обозначение" (повторный заголовок)
        if designation and "обозначение" in str(designation).lower():
            continue
        
        # Конвертируем кол-во
        try:
            qty = float(qty_raw) if qty_raw is not None else 0.0
        except (ValueError, TypeError):
            qty = 0.0
        
        # Конвертируем номер позиции
        try:
            pos_number = int(float(str(pos_raw))) if pos_raw is not None else None
        except (ValueError, TypeError):
            pos_number = None
        
        positions.append({
            "pos_number": pos_number,
            "designation": str(designation).strip() if designation else "",
            "name": str(name).strip() if name else "",
            "qty": qty
        })
    
    if not positions:
        raise ValueError("В файле не найдено ни одной позиции")
    
    # Убираем дубликаты (одинаковое обозначение + наименование подряд — артефакт мёрджа)
    seen = set()
    unique_positions = []
    for p in positions:
        key = (p["designation"], p["name"])
        if key not in seen:
            seen.add(key)
            unique_positions.append(p)
    
    return {
        "order_number": order_number,
        "positions": unique_positions
    }
